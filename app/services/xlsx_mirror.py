"""Mirror writes to Connor's existing Baseball_Cards.xlsx workbook.

We append rows under the Inventory tab using the existing column layout.
This is a failsafe so the data lives in the same workbook he's been
maintaining, even if Google Sheets is offline.

Sales Log tab structure (header at row 4):
  A: Month, B: Individual Cards Sold (count), C: Bulk Lots Sold (count),
  D: Gross Revenue, E: Fees Total, F: Net (=D-E formula), G: Notes
  Rows 5..N are monthly buckets (e.g. "May 2026").
  Row 13 (or last data row+1) is a TOTAL row.
  We UPDATE the matching month row rather than appending new rows.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional
from datetime import datetime

import openpyxl

from app.config import settings
from app import models


# These column indices match the existing Inventory header row (row 4):
# A: Card # (formula), B: Year, C: Set, D: Player, E: Card # in Set,
# F: Parallel, G: Condition, H: Storage, I: Est Raw, J: Est Graded,
# K: Comps URL, L: Status, M: Channel, N: Sold Price, O: Fee %,
# P: Net (formula), Q: Notes
HEADER_ROW = 4

# Sales Log header is at row 4; monthly data starts at row 5
SALES_LOG_HEADER_ROW = 4
SALES_LOG_DATA_START = 5


def _open():
    path = settings.local_xlsx_path
    if not path or not Path(path).exists():
        return None, None
    wb = openpyxl.load_workbook(path)
    if "Inventory" not in wb.sheetnames:
        return None, None
    return wb, wb["Inventory"]


def _open_full() -> Optional[openpyxl.Workbook]:
    """Open the workbook (all sheets). Returns None if not configured/found."""
    path = settings.local_xlsx_path
    if not path or not Path(path).exists():
        return None
    return openpyxl.load_workbook(path)


def append_card(card: models.Card) -> bool:
    wb, ws = _open()
    if ws is None:
        return False
    # find first empty row after header
    row = HEADER_ROW + 1
    while ws.cell(row=row, column=2).value is not None:
        row += 1
    ws.cell(row=row, column=1, value="=ROW()-4")
    ws.cell(row=row, column=2, value=card.year)
    ws.cell(row=row, column=3, value=card.set_brand)
    ws.cell(row=row, column=4, value=card.player)
    ws.cell(row=row, column=5, value=card.card_no)
    ws.cell(row=row, column=6, value=card.parallel)
    ws.cell(row=row, column=7, value=card.condition)
    ws.cell(row=row, column=8, value=card.storage)
    ws.cell(row=row, column=9, value=card.est_value_raw or card.comp_median_weighted or card.comp_median)
    ws.cell(row=row, column=10, value=card.est_value_graded)
    ws.cell(row=row, column=11, value=card.comp_url)
    ws.cell(row=row, column=12, value=card.status)
    ws.cell(row=row, column=13, value=card.channel)
    ws.cell(row=row, column=14, value=card.sold_price)
    ws.cell(row=row, column=15, value=card.fee_pct or 0.13)
    ws.cell(row=row, column=16, value=f"=IFERROR(N{row}*(1-O{row}),0)")
    notes_bits = []
    if card.notes:
        notes_bits.append(card.notes)
    if card.is_hit_watchlist and card.hit_reason:
        notes_bits.append(f"HIT: {card.hit_reason}")
    notes_bits.append(f"Auto-cataloged {datetime.utcnow():%Y-%m-%d}")
    ws.cell(row=row, column=17, value=" | ".join(notes_bits))
    wb.save(settings.local_xlsx_path)
    return True


def update_sales_log(card: models.Card) -> bool:
    """Update the Sales Log monthly roll-up when a Card transitions to Sold.

    The Sales Log tab has monthly rows (e.g. "May 2026") with columns:
      A: Month, B: Cards Sold count, C: Bulk Lots count,
      D: Gross Revenue, E: Fees Total, F: Net (formula), G: Notes

    We find the row whose Month matches the sale month and increment
    the count + revenue/fees in place. If no matching month row exists,
    we insert a new one before the TOTAL row.
    """
    wb = _open_full()
    if wb is None:
        return False
    if "Sales Log" not in wb.sheetnames:
        return False

    ws = wb["Sales Log"]

    sold_at = card.sold_at or datetime.utcnow()
    month_label = sold_at.strftime("%b %Y")  # e.g. "May 2026"
    gross = card.sold_price or 0.0
    fees = round(gross * (card.fee_pct or 0.13), 2)

    # Scan data rows for matching month (stop at TOTAL or empty month col)
    match_row = None
    total_row = None
    last_data_row = SALES_LOG_DATA_START - 1

    for row_idx in range(SALES_LOG_DATA_START, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            break
        cell_str = str(cell_val).strip()
        if cell_str.upper() == "TOTAL":
            total_row = row_idx
            break
        if cell_str == month_label:
            match_row = row_idx
        last_data_row = row_idx

    if match_row is not None:
        # Increment existing row
        cur_count = ws.cell(row=match_row, column=2).value or 0
        cur_gross = ws.cell(row=match_row, column=4).value or 0.0
        cur_fees = ws.cell(row=match_row, column=5).value or 0.0
        ws.cell(row=match_row, column=2, value=int(cur_count) + 1)
        ws.cell(row=match_row, column=4, value=round(float(cur_gross) + gross, 2))
        ws.cell(row=match_row, column=5, value=round(float(cur_fees) + fees, 2))
        # Preserve the Net formula if it exists; otherwise write it
        f_cell = ws.cell(row=match_row, column=6)
        if not str(f_cell.value or "").startswith("="):
            f_cell.value = f"=D{match_row}-E{match_row}"
    else:
        # Insert a new month row before the TOTAL row (or append after last data row)
        insert_at = (total_row if total_row is not None else last_data_row + 1)
        if total_row is not None:
            ws.insert_rows(insert_at)
            # After insert, total_row is now one lower — formulas referencing that
            # range will auto-adjust because openpyxl shifts them.
        row_idx = insert_at
        ws.cell(row=row_idx, column=1, value=month_label)
        ws.cell(row=row_idx, column=2, value=1)          # cards sold count
        ws.cell(row=row_idx, column=3, value=0)          # bulk lots
        ws.cell(row=row_idx, column=4, value=gross)
        ws.cell(row=row_idx, column=5, value=fees)
        ws.cell(row=row_idx, column=6, value=f"=D{row_idx}-E{row_idx}")

    wb.save(settings.local_xlsx_path)
    return True
