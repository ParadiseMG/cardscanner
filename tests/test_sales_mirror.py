"""A5: Tests for xlsx Sales Log mirror on Card sold transition."""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from app import models
from app.services import xlsx_mirror


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def real_xlsx(tmp_path) -> Path:
    """Copy the real Baseball_Cards.xlsx into tmp_path for isolated testing."""
    src = Path("/sessions/nifty-dreamy-fermi/mnt/Documents/Personal Finances/Baseball_Cards.xlsx")
    dst = tmp_path / "Baseball_Cards.xlsx"
    if src.exists():
        shutil.copy2(src, dst)
    else:
        # Build a minimal workbook that matches the Sales Log schema
        wb = openpyxl.Workbook()
        # Inventory sheet (required by xlsx_mirror._open)
        inv = wb.active
        inv.title = "Inventory"
        inv.cell(row=4, column=1, value="Card #")
        # Sales Log sheet
        sl = wb.create_sheet("Sales Log")
        sl.cell(row=1, column=1, value="Sales Log — monthly roll-up")
        sl.cell(row=4, column=1, value="Month")
        sl.cell(row=4, column=2, value="Individual Cards Sold (count)")
        sl.cell(row=4, column=3, value="Bulk Lots Sold (count)")
        sl.cell(row=4, column=4, value="Gross Revenue")
        sl.cell(row=4, column=5, value="Fees Total")
        sl.cell(row=4, column=6, value="Net")
        sl.cell(row=4, column=7, value="Notes")
        # Prepopulate a month row
        sl.cell(row=5, column=1, value="May 2026")
        sl.cell(row=5, column=2, value=0)
        sl.cell(row=5, column=3, value=0)
        sl.cell(row=5, column=4, value=0)
        sl.cell(row=5, column=5, value=0)
        sl.cell(row=5, column=6, value="=D5-E5")
        # TOTAL row
        sl.cell(row=6, column=1, value="TOTAL")
        sl.cell(row=6, column=2, value="=SUM(B5:B5)")
        sl.cell(row=6, column=3, value="=SUM(C5:C5)")
        sl.cell(row=6, column=4, value="=SUM(D5:D5)")
        sl.cell(row=6, column=5, value="=SUM(E5:E5)")
        sl.cell(row=6, column=6, value="=SUM(F5:F5)")
        wb.save(dst)
    return dst


def _make_sold_card(sold_at: datetime = None, sold_price: float = 25.0,
                    fee_pct: float = 0.13) -> models.Card:
    return models.Card(
        id=1,
        player="Ken Griffey Jr",
        year=1989,
        set_brand="Upper Deck",
        status="Sold",
        sold_price=sold_price,
        fee_pct=fee_pct,
        sold_at=sold_at or datetime(2026, 5, 10, 12, 0, 0),
        sale_channel="eBay",
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_update_sales_log_increments_matching_month(real_xlsx, monkeypatch):
    """Selling a card in May 2026 increments the May 2026 row."""
    monkeypatch.setattr(xlsx_mirror.settings, "local_xlsx_path", str(real_xlsx))

    card = _make_sold_card(sold_at=datetime(2026, 5, 10), sold_price=25.0, fee_pct=0.13)
    result = xlsx_mirror.update_sales_log(card)
    assert result is True

    wb = openpyxl.load_workbook(real_xlsx, data_only=True)
    sl = wb["Sales Log"]

    # Find May 2026 row
    may_row = None
    for row_idx in range(5, sl.max_row + 1):
        if str(sl.cell(row=row_idx, column=1).value or "").strip() == "May 2026":
            may_row = row_idx
            break

    assert may_row is not None, "May 2026 row not found"
    assert sl.cell(row=may_row, column=2).value == 1       # count = 1
    assert sl.cell(row=may_row, column=4).value == pytest.approx(25.0)  # gross
    assert sl.cell(row=may_row, column=5).value == pytest.approx(3.25, abs=0.01)  # fees


def test_update_sales_log_accumulates_multiple_sales(real_xlsx, monkeypatch):
    """Two May 2026 sales accumulate in the same row."""
    monkeypatch.setattr(xlsx_mirror.settings, "local_xlsx_path", str(real_xlsx))

    card1 = _make_sold_card(sold_at=datetime(2026, 5, 10), sold_price=20.0)
    card2 = _make_sold_card(sold_at=datetime(2026, 5, 15), sold_price=30.0)

    xlsx_mirror.update_sales_log(card1)
    xlsx_mirror.update_sales_log(card2)

    wb = openpyxl.load_workbook(real_xlsx, data_only=True)
    sl = wb["Sales Log"]

    may_row = None
    for row_idx in range(5, sl.max_row + 1):
        if str(sl.cell(row=row_idx, column=1).value or "").strip() == "May 2026":
            may_row = row_idx
            break

    assert may_row is not None
    assert sl.cell(row=may_row, column=2).value == 2       # two sales
    assert sl.cell(row=may_row, column=4).value == pytest.approx(50.0)  # gross total


def test_update_sales_log_new_month_inserted(real_xlsx, monkeypatch):
    """A card sold in a month not in the log inserts a new row before TOTAL."""
    monkeypatch.setattr(xlsx_mirror.settings, "local_xlsx_path", str(real_xlsx))

    # Sell a card in Jun 2026 (not pre-populated in the minimal fixture)
    card = _make_sold_card(sold_at=datetime(2026, 6, 5), sold_price=40.0)
    result = xlsx_mirror.update_sales_log(card)
    assert result is True

    wb = openpyxl.load_workbook(real_xlsx, data_only=True)
    sl = wb["Sales Log"]

    jun_row = None
    for row_idx in range(5, sl.max_row + 1):
        if str(sl.cell(row=row_idx, column=1).value or "").strip() == "Jun 2026":
            jun_row = row_idx
            break

    # If the real xlsx already has a Jun 2026 row it will be found; otherwise
    # we inserted one.
    assert jun_row is not None, "Jun 2026 row not found after insert"
    assert sl.cell(row=jun_row, column=2).value == 1


def test_update_sales_log_no_path(monkeypatch):
    """Returns False gracefully when xlsx path is not configured."""
    monkeypatch.setattr(xlsx_mirror.settings, "local_xlsx_path", "")
    card = _make_sold_card()
    result = xlsx_mirror.update_sales_log(card)
    assert result is False


def test_update_sales_log_missing_file(tmp_path, monkeypatch):
    """Returns False gracefully when the xlsx file does not exist."""
    monkeypatch.setattr(xlsx_mirror.settings, "local_xlsx_path",
                        str(tmp_path / "nonexistent.xlsx"))
    card = _make_sold_card()
    result = xlsx_mirror.update_sales_log(card)
    assert result is False


def test_update_sales_log_net_formula_preserved(real_xlsx, monkeypatch):
    """Net column retains a formula, not a hardcoded value."""
    monkeypatch.setattr(xlsx_mirror.settings, "local_xlsx_path", str(real_xlsx))

    card = _make_sold_card(sold_at=datetime(2026, 5, 10), sold_price=25.0)
    xlsx_mirror.update_sales_log(card)

    # Load with data_only=False to see formulas
    wb = openpyxl.load_workbook(real_xlsx, data_only=False)
    sl = wb["Sales Log"]

    may_row = None
    for row_idx in range(5, sl.max_row + 1):
        if str(sl.cell(row=row_idx, column=1).value or "").strip() == "May 2026":
            may_row = row_idx
            break

    assert may_row is not None
    f_cell_val = sl.cell(row=may_row, column=6).value
    assert str(f_cell_val or "").startswith("="), f"Expected formula, got: {f_cell_val!r}"
